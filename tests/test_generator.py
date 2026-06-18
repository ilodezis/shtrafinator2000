import os
import datetime
import zipfile
import pytest
from lxml import etree

import generator

def test_decline_to_dative():
    # Test male names ending in various consonants
    assert generator.decline_to_dative("Завгороднев Григорий Викторович") == "Завгородневу Григорию Викторовичу"
    assert generator.decline_to_dative("Романюк Максим Юрьевич") == "Романюку Максиму Юрьевичу"
    assert generator.decline_to_dative("Толкачёв Виктор Григорьевич") == "Толкачёву Виктору Григорьевичу"
    assert generator.decline_to_dative("Фомин Сергей Петрович") == "Фомину Сергею Петровичу"
    
    # Test male names ending in -ый / -ий / -ой
    assert generator.decline_to_dative("Нагорный Роман Дмитриевич") == "Нагорному Роману Дмитриевичу"
    
    # Test names ending in -ко (indeclinable)
    assert generator.decline_to_dative("Боваренко Сергей Владимирович") == "Боваренко Сергею Владимировичу"
    assert generator.decline_to_dative("Онипко Анна Витальевна") == "Онипко Анне Витальевне"
    
    # Test female names ending in -а / -я
    assert generator.decline_to_dative("Алехина Елена Сергеевна") == "Алехиной Елене Сергеевне"
    assert generator.decline_to_dative("Губанихина Юлия Николаевна") == "Губанихиной Юлии Николаевне"
    assert generator.decline_to_dative("Садовникова Полина Витальевна") == "Садовниковой Полине Витальевне"

def test_derive_director():
    # IP name should be declined
    name, missing = generator.derive_director({"yl": "ИП Завгороднев Григорий Викторович"})
    assert name == "Завгородневу Григорию Викторовичу"
    assert not missing

    # LLC/ООО name should return placeholder when director is not explicitly given
    name, missing = generator.derive_director({"yl": "ООО \"Атлантик Айти\""})
    assert name == "[Руководитель не указан]"
    assert missing

    # Explicit director in record should be used and declined
    name, missing = generator.derive_director({"yl": "ООО \"Атлантик Айти\"", "director": "Завгороднев Григорий Викторович"})
    assert name == "Завгородневу Григорию Викторовичу"
    assert not missing

def test_document_protection_removed():
    # Generate a sample record with custom columns
    record = {
        "period": datetime.datetime(2026, 2, 1),
        "inn": "344708426365",
        "yl": "ООО \"Атлантик Айти\"",
        "fine": 30000.0,
        "fraud_pct": 0.0186,
        "director": "Боваренко Сергей Владимирович",
        "ogrn": "1234567890123",
        "email": "test@atlantic.it"
    }
    doc_bytes, warnings = generator.fill_template(record, datetime.datetime.now())
    
    # Warnings for missing director, ogrn or email should NOT be present
    assert "Руководитель не указан" not in warnings
    assert "ОГРН не указан" not in warnings
    assert "Email не указан" not in warnings

    # Verify the generated zip does not have document protection in settings.xml
    import io
    zin = zipfile.ZipFile(io.BytesIO(doc_bytes), 'r')
    settings_xml = zin.read('word/settings.xml')
    root = etree.fromstring(settings_xml)
    
    W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    dp = root.findall(W + "documentProtection")
    assert len(dp) == 0, "documentProtection tag was not removed from settings.xml"
    
    # Verify values inside document.xml
    doc_xml = zin.read('word/document.xml')
    doc_root = etree.fromstring(doc_xml)
    doc_text = "".join(doc_root.itertext())
    assert "Боваренко Сергею Владимировичу" in doc_text
    assert "1234567890123" in doc_text
    assert "test@atlantic.it" in doc_text
    
    zin.close()


def test_header_matches():
    # Test AND logic
    assert generator._header_matches("Отчетный период", [["отчетный", "период"]])
    assert not generator._header_matches("Период", [["отчетный", "период"]])
    
    # Test OR logic
    assert generator._header_matches("директор", [["директор"], ["руководитель"]])
    assert generator._header_matches("руководитель", [["директор"], ["руководитель"]])
    assert not generator._header_matches("менеджер", [["директор"], ["руководитель"]])


def test_read_excel_optional_columns(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Для соп-я"
    
    # Write headers including optional ones
    headers = [
        "Отчетный период", "ИНН", "ЮЛ", "Оборот, руб.", "Фрод, руб.",
        "Штраф", "Решение (да/нет)", "% фрода от оборота",
        "Гендир", "ОГРН", "Почта"
    ]
    ws.append(headers)
    
    # Write data row
    ws.append([
        "2026-02-01", 1234567890, "ООО Тест", 100000, 1000,
        5000, "да", 0.01,
        "Иванов Иван Иванович", 1027700123456, "test@test.ru"
    ])
    
    excel_file = tmp_path / "test.xlsx"
    wb.save(excel_file)
    
    records = generator.read_excel(str(excel_file))
    assert len(records) == 1
    rec = records[0]
    assert rec["yl"] == "ООО Тест"
    assert rec["director"] == "Иванов Иван Иванович"
    assert rec["ogrn"] == "1027700123456"
    assert rec["email"] == "test@test.ru"


def test_read_excel_missing_optional_columns(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Для соп-я"
    
    # Write headers, omitting optional ones (director, ogrn, email)
    headers = [
        "Отчетный период", "ИНН", "ЮЛ", "Оборот, руб.", "Фрод, руб.",
        "Штраф", "Решение (да/нет)", "% фрода от оборота"
    ]
    ws.append(headers)
    
    ws.append([
        "2026-02-01", 1234567890, "ООО Тест", 100000, 1000,
        5000, "да", 0.01
    ])
    
    excel_file = tmp_path / "test.xlsx"
    wb.save(excel_file)
    
    # Verify it loads without error (skips/passes missing columns)
    records = generator.read_excel(str(excel_file))
    assert len(records) == 1
    rec = records[0]
    assert rec["yl"] == "ООО Тест"
    assert rec["director"] is None
    assert rec["ogrn"] is None
    assert rec["email"] is None


def test_coerce_date():
    # Настоящая дата проходит как есть
    dt = datetime.datetime(2026, 2, 1)
    assert generator._coerce_date(dt) == dt
    # datetime.date нормализуется в datetime
    assert generator._coerce_date(datetime.date(2026, 2, 1)) == datetime.datetime(2026, 2, 1)
    # Excel-серийник (1900 date system)
    assert generator._coerce_date(46054) == datetime.datetime(2026, 2, 1)
    # Текстовые форматы
    assert generator._coerce_date("2026-02-01") == datetime.datetime(2026, 2, 1)
    assert generator._coerce_date("01.02.2026") == datetime.datetime(2026, 2, 1)
    # Нераспознаваемое -> None (а не падение)
    assert generator._coerce_date("февраль 2026") is None
    assert generator._coerce_date(None) is None
    assert generator._coerce_date(True) is None


def test_fill_template_unparseable_period_placeholder():
    import io
    record = {
        "period": None,  # период не распознан read_excel
        "inn": "1234567890",
        "yl": "ООО \"Тест\"",
        "fine": 10000.0,
        "fraud_pct": 0.01,
        "director": "Иванов Иван Иванович",
        "ogrn": "1027700123456",
        "email": "test@test.ru",
    }
    doc_bytes, warnings = generator.fill_template(record, datetime.datetime.now())
    assert "Период не распознан" in warnings
    doc_text = "".join(etree.fromstring(zipfile.ZipFile(io.BytesIO(doc_bytes)).read("word/document.xml")).itertext())
    assert "[Период не указан]" in doc_text


def test_read_excel_textual_period(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Для соп-я"
    headers = [
        "Отчетный период", "ИНН", "ЮЛ", "Фрод, руб.",
        "Штраф", "Решение (да/нет)", "% фрода от оборота",
    ]
    ws.append(headers)
    # Период как текст в распознаваемом формате
    ws.append(["01.02.2026", 1234567890, "ООО Тест", 1000, 5000, "да", 0.01])
    # Период как нераспознаваемый текст -> None, но строка не теряется
    ws.append(["неизвестно", 1234567890, "ООО Тест2", 1000, 5000, "да", 0.01])
    excel_file = tmp_path / "test.xlsx"
    wb.save(excel_file)

    records = generator.read_excel(str(excel_file))
    assert len(records) == 2
    assert records[0]["period"] == datetime.datetime(2026, 2, 1)
    assert records[1]["period"] is None


def test_custom_signatory_injection():
    import io
    record = {
        "period": datetime.datetime(2026, 2, 1),
        "inn": "1234567890",
        "yl": "ООО \"Тест\"",
        "fine": 10000.0,
        "fraud_pct": 0.01,
        "director": "Иванов Иван Иванович",
        "ogrn": "1027700123456",
        "email": "test@test.ru"
    }
    
    # 1. Test default signatory (empty)
    doc_bytes, _ = generator.fill_template(record, datetime.datetime.now())
    zin = zipfile.ZipFile(io.BytesIO(doc_bytes), 'r')
    doc_text = "".join(etree.fromstring(zin.read('word/document.xml')).itertext())
    assert "____________/[Подписант не указан]" in doc_text
    zin.close()
    
    # 2. Test custom signatory
    doc_bytes, _ = generator.fill_template(record, datetime.datetime.now(), signatory="Петров П.П.")
    zin = zipfile.ZipFile(io.BytesIO(doc_bytes), 'r')
    doc_text = "".join(etree.fromstring(zin.read('word/document.xml')).itertext())
    assert "____________/Петров П.П." in doc_text
    zin.close()
