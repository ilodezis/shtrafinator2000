import base64
import datetime
import io
import os
import re
import zipfile
from typing import Any, Callable, Optional

import openpyxl
from lxml import etree

import template_data

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

MONTHS_NOM = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}

HEADER_ANCHORS = {
    "period": [["отчетный", "период"]],
    "inn": [["инн"]],
    "yl": [["юл"]],
    "turnover": [["оборот"]],
    "fraud_rub": [["фрод", "руб"]],
    "fine": [["штраф"]],
    "decision": [["решение"]],
    "fraud_pct": [["% фрод"], ["фрода от оборота"]],
    "director": [["директор"], ["руководитель"], ["гендир"], ["фио"]],
    "ogrn": [["огрн"]],
    "email": [["email"], ["e-mail"], ["почта"]],
}


def _normalize(text: str) -> str:
    return text.replace("\xa0", " ").strip().lower()


def _header_matches(header: str, options: list[list[str]]) -> bool:
    h = _normalize(header)
    return any(all(word in h for word in option) for option in options)


def _find_columns(ws: Any) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            continue
        header = str(val)
        for field, anchors in HEADER_ANCHORS.items():
            if field not in mapping and _header_matches(header, anchors):
                mapping[field] = col
                break
    missing = set(HEADER_ANCHORS.keys()) - set(mapping.keys()) - {"turnover", "director", "ogrn", "email"}
    if missing:
        raise ValueError(
            f"В заголовках не найдены колонки: {', '.join(sorted(missing))}"
        )
    return mapping


class GenerationResult:
    def __init__(self, filename: str, success: bool,
                 warnings: list[str] | None = None,
                 error: str | None = None):
        self.filename = filename
        self.success = success
        self.warnings = warnings or []
        self.error = error


def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name == "Для соп-я":
            sheet_name = name
            break
    if sheet_name is None:
        for name in wb.sheetnames:
            if "соп" in name.lower():
                sheet_name = name
                break
    if sheet_name is None:
        raise ValueError(
            f"Лист «Для соп-я» не найден. Доступные листы: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    cols = _find_columns(ws)

    records: list[dict] = []
    for row in range(2, ws.max_row + 1):
        decision = ws.cell(row=row, column=cols["decision"]).value
        if decision is None or str(decision).strip().lower() != "да":
            continue

        period = ws.cell(row=row, column=cols["period"]).value
        inn_val = ws.cell(row=row, column=cols["inn"]).value
        yl = ws.cell(row=row, column=cols["yl"]).value
        fine = ws.cell(row=row, column=cols["fine"]).value
        fraud_pct = ws.cell(row=row, column=cols["fraud_pct"]).value

        if any(v is None for v in [period, inn_val, yl, fine]):
            continue

        if isinstance(period, (int, float)):
            period = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(period))

        director_val = ws.cell(row=row, column=cols["director"]).value if "director" in cols else None
        ogrn_val = ws.cell(row=row, column=cols["ogrn"]).value if "ogrn" in cols else None
        email_val = ws.cell(row=row, column=cols["email"]).value if "email" in cols else None

        records.append({
            "period": period,
            "inn": str(int(inn_val)) if isinstance(inn_val, (int, float)) else str(inn_val),
            "yl": str(yl).strip(),
            "fine": float(fine),
            "fraud_pct": float(fraud_pct) if fraud_pct is not None else None,
            "row_num": row,
            "director": str(director_val).strip() if director_val is not None else None,
            "ogrn": str(int(ogrn_val)) if isinstance(ogrn_val, (int, float)) else (str(ogrn_val).strip() if ogrn_val is not None else None),
            "email": str(email_val).strip() if email_val is not None else None,
        })

    return records


def format_fine(value: float) -> str:
    rounded = round(value)
    formatted = f"{rounded:,}".replace(",", " ")
    return f"{formatted} руб."


def format_fraud_pct(value: float) -> str:
    pct = value * 100
    return f"{pct:.2f}".replace(".", ",")


def format_period(dt: datetime.datetime) -> str:
    return f"{MONTHS_NOM[dt.month]} {dt.year}"


def format_date(dt: datetime.datetime) -> str:
    return f"б/н от {dt.strftime('%d.%m.%Y')}г."


def decline_to_dative(name_str: str) -> str:
    parts = [p.strip() for p in name_str.split() if p.strip()]
    if len(parts) < 2:
        return name_str
    
    lastname = parts[0]
    firstname = parts[1]
    patronymic = parts[2] if len(parts) > 2 else ""
    
    is_male = True
    if patronymic:
        if patronymic.endswith("на"):
            is_male = False
        elif patronymic.endswith("ич"):
            is_male = True
    else:
        if firstname.endswith(("а", "я", "и")):
            is_male = False
            
    # Decline LastName
    dative_lastname = lastname
    if is_male:
        if lastname.endswith("ко"):
            pass
        elif lastname.endswith("ый") or lastname.endswith("ий"):
            dative_lastname = lastname[:-2] + "ому"
        elif lastname.endswith("ой"):
            dative_lastname = lastname[:-2] + "ому"
        elif lastname.endswith(("в", "н", "л", "р", "к", "ч", "м", "т", "с", "х", "ц", "г", "д", "з", "б", "п", "ф")):
            dative_lastname = lastname + "у"
        elif lastname.endswith("й"):
            dative_lastname = lastname[:-1] + "ю"
        else:
            dative_lastname = lastname + "у"
    else:
        if lastname.endswith("ко"):
            pass
        elif lastname.endswith("а") or lastname.endswith("я"):
            if lastname.endswith("ая"):
                dative_lastname = lastname[:-2] + "ой"
            else:
                dative_lastname = lastname[:-1] + "ой"

    # Decline FirstName
    dative_firstname = firstname
    if is_male:
        if firstname.endswith("й"):
            dative_firstname = firstname[:-1] + "ю"
        elif firstname.endswith("ь"):
            dative_firstname = firstname[:-1] + "ю"
        elif firstname.endswith(("а", "я")):
            dative_firstname = firstname[:-1] + "е"
        else:
            dative_firstname = firstname + "у"
    else:
        if firstname.endswith("ия"):
            dative_firstname = firstname[:-1] + "и"
        elif firstname.endswith(("а", "я")):
            dative_firstname = firstname[:-1] + "е"
        elif firstname.endswith("ь"):
            dative_firstname = firstname[:-1] + "и"

    # Decline Patronymic
    dative_patronymic = patronymic
    if patronymic:
        if is_male:
            if patronymic.endswith("ич"):
                dative_patronymic = patronymic + "у"
        else:
            if patronymic.endswith("на"):
                dative_patronymic = patronymic[:-1] + "е"
                
    if dative_patronymic:
        return f"{dative_lastname} {dative_firstname} {dative_patronymic}"
    else:
        return f"{dative_lastname} {dative_firstname}"



def derive_director(record: dict) -> tuple[str, bool]:
    # 1. If explicit director is in the excel record
    if record.get("director"):
        director_name = record["director"].strip()
        if director_name:
            return decline_to_dative(director_name), False
            
    # 2. Fallback to extracting from ИП name
    stripped = record["yl"].strip()
    if stripped.upper().startswith("ИП "):
        director_name = stripped[3:].strip()
        return decline_to_dative(director_name), False
        
    return "[Руководитель не указан]", True



def sanitize_filename(yl: str) -> str:
    clean = yl
    for ch in '«»""\'/\\:*?<>|':
        clean = clean.replace(ch, "_")
    clean = re.sub(r"_+", "_", clean).strip("_")
    return f"{clean}.docx"


def _resolve_duplicate(filepath: str, used: set[str]) -> str:
    if filepath.lower() not in used:
        used.add(filepath.lower())
        return filepath
    base, ext = os.path.splitext(filepath)
    counter = 2
    while f"{base}_{counter}{ext}".lower() in used:
        counter += 1
    result = f"{base}_{counter}{ext}"
    used.add(result.lower())
    return result


def _make_run(text: str, bold: bool = False) -> etree._Element:
    r = etree.Element(W + "r")
    rpr = etree.SubElement(r, W + "rPr")
    fonts = etree.SubElement(rpr, W + "rFonts")
    fonts.set(W + "ascii", "Candara")
    fonts.set(W + "hAnsi", "Candara")
    fonts.set(W + "cs", "Times New Roman")
    if bold:
        etree.SubElement(rpr, W + "b")
    sz = etree.SubElement(rpr, W + "sz")
    sz.set(W + "val", "24")
    szCs = etree.SubElement(rpr, W + "szCs")
    szCs.set(W + "val", "24")
    t = etree.SubElement(r, W + "t")
    t.text = text
    if text and (text[0] == " " or text[-1] == " "):
        t.set(XML_SPACE, "preserve")
    return r


def _make_paragraph(text: str, bold: bool = False) -> etree._Element:
    p = etree.Element(W + "p")
    ppr = etree.SubElement(p, W + "pPr")
    rpr = etree.SubElement(ppr, W + "rPr")
    fonts = etree.SubElement(rpr, W + "rFonts")
    fonts.set(W + "ascii", "Candara")
    fonts.set(W + "hAnsi", "Candara")
    fonts.set(W + "cs", "Times New Roman")
    if bold:
        etree.SubElement(rpr, W + "b")
    sz = etree.SubElement(rpr, W + "sz")
    sz.set(W + "val", "24")
    szCs = etree.SubElement(rpr, W + "szCs")
    szCs.set(W + "val", "24")
    r = _make_run(text, bold=bold)
    p.append(r)
    return p


def _replace_sdt(sdt: etree._Element, replacement: etree._Element) -> None:
    parent = sdt.getparent()
    idx = list(parent).index(sdt)
    parent.insert(idx, replacement)
    parent.remove(sdt)


def fill_template(record: dict, letter_date: datetime.datetime, signatory: str = "Жаворонкина А.М.") -> tuple[bytes, list[str]]:
    warnings: list[str] = []

    director, director_missing = derive_director(record)
    if director_missing:
        warnings.append("Руководитель не указан")

    inn = record["inn"]
    yl = record["yl"]
    if record["fraud_pct"] is not None:
        fraud_pct_str = format_fraud_pct(record["fraud_pct"])
    else:
        fraud_pct_str = "[% не указан]"
        warnings.append("% фрода не указан")
    period_str = format_period(record["period"])
    fine_str = format_fine(record["fine"])
    date_str = format_date(letter_date)
    
    # OGRN
    if record.get("ogrn"):
        ogrn_str = record["ogrn"]
    else:
        ogrn_str = "[ОГРН не указан]"
        warnings.append("ОГРН не указан")
        
    # Email
    if record.get("email"):
        email_str = record["email"]
    else:
        email_str = "[Email не указан]"
        warnings.append("Email не указан")

    template_bytes = template_data.get_template_bytes()
    src = io.BytesIO(template_bytes)
    zin = zipfile.ZipFile(src, "r")
    xml_bytes = zin.read("word/document.xml")
    root = etree.fromstring(xml_bytes)

    sdts = list(root.iter(W + "sdt"))
    if len(sdts) != 12:
        warnings.append(f"Ожидалось 12 SDT, найдено {len(sdts)} — результат может быть некорректным")

    # SDT mapping (0-indexed from iter order):
    #  0: date field (inline, no placeholder)
    #  1: director (block-level, placeholder)
    #  2: org name header (block-level, placeholder)
    #  3: ОГРН (inline, placeholder)
    #  4: ИНН (inline, placeholder)
    #  5: email outer (inline, no placeholder — wraps #6)
    #  6: email inner (nested, placeholder)
    #  7: org name body (inline, placeholder)
    #  8: fraud % (inline, placeholder)
    #  9: period (inline, placeholder)
    # 10: fine (inline, placeholder)
    # 11: signature (inline, no placeholder — skip)

    replacements = {
        0: ("inline", date_str, False),
        1: ("block", director, True),
        2: ("block", yl, True),
        3: ("inline", ogrn_str, False),
        4: ("inline", inn, False),
        5: ("inline", email_str, False),
        # 6 is removed when 5 is replaced
        7: ("inline", yl, False),
        8: ("inline", fraud_pct_str, False),
        9: ("inline", period_str, False),
        10: ("inline", fine_str, False),
        11: ("inline", f" ____________/{signatory}", False),
    }

    for idx in sorted(replacements.keys(), reverse=True):
        if idx >= len(sdts):
            continue
        mode, value, bold = replacements[idx]
        sdt = sdts[idx]
        if mode == "block":
            _replace_sdt(sdt, _make_paragraph(value, bold=bold))
        else:
            _replace_sdt(sdt, _make_run(value, bold=bold))

    modified_xml = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "word/document.xml":
                zout.writestr(item, modified_xml)
            elif item.filename == "word/settings.xml":
                settings_bytes = zin.read(item.filename)
                try:
                    s_root = etree.fromstring(settings_bytes)
                    for dp in s_root.findall(W + "documentProtection"):
                        s_root.remove(dp)
                    modified_settings = etree.tostring(s_root, xml_declaration=True, encoding="UTF-8", standalone=True)
                    zout.writestr(item, modified_settings)
                except Exception:
                    zout.writestr(item, settings_bytes)
            else:
                zout.writestr(item, zin.read(item.filename))
    zin.close()

    return out.getvalue(), warnings


def generate_all(
    excel_path: str,
    output_dir: str,
    letter_date: datetime.datetime,
    signatory: str = "Жаворонкина А.М.",
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> list[GenerationResult]:

    records = read_excel(excel_path)
    if not records:
        raise ValueError("Не найдено строк с решением «да» для обработки")

    results: list[GenerationResult] = []
    used_filenames: set[str] = set()
    total = len(records)

    for i, record in enumerate(records):
        filename = sanitize_filename(record["yl"])
        filepath = os.path.join(output_dir, filename)
        filepath = _resolve_duplicate(filepath, used_filenames)
        actual_filename = os.path.basename(filepath)

        try:
            doc_bytes, warnings = fill_template(record, letter_date, signatory)
            with open(filepath, "wb") as f:
                f.write(doc_bytes)
            results.append(GenerationResult(actual_filename, True, warnings=warnings))
        except Exception as e:
            results.append(GenerationResult(actual_filename, False, error=str(e)))

        if progress_callback:
            progress_callback(i + 1, total)

    return results
